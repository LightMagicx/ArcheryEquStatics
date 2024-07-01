from enum import Enum

import openpyxl

class Archer:
    def __init__(self,name,age,ranking,gender,country,\
                 riser,limb,sightbar,sightpin,stabilizer,tab):
        self.name=name
        self.age=age if age else 'N/A'
        self.ranking=ranking
        self.gender=gender
        self.country=country
        self.riser=riser
        self.limb=limb if limb else 'N/A'
        self.sightbar=sightbar if sightbar else 'N/A'
        self.sightpin=sightpin if sightpin else 'N/A'
        self.stabilizer=stabilizer if stabilizer else 'N/A'
        self.tab=tab if tab else 'N/A'

    def displayInfo(self):
        #
        print("{:<25} {:<15} {:<15} {:<15} {:<15}".\
              format(self.name,self.age,self.ranking,self.gender,self.country))

    def displayEqu(self):
        print("{:<25} {:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".\
              format(self.name,self.riser,self.limb,self.sightbar,self.sightpin,self.stabilizer,self.tab))
        
class ArcherList:
    def __init__(self):
        self.aList=[]

    def loadDatas(self,filepath):
        workbook=openpyxl.load_workbook(filepath)
        sheet_names=workbook.sheetnames

        for s_name in sheet_names:
            sheet=workbook[s_name]
            print(f'Loading sheet: {s_name}')

            for row in range(2,sheet.max_row):
                temp=[]
                for cell in sheet[row]:
                    temp.append(cell.value)
                if temp[5] != None and temp[0]!=None:
                    archer=Archer(temp[0],temp[1],temp[2],temp[3],temp[4],temp[5],temp[6],temp[7],temp[8],temp[9],temp[10])
                    #archer.displayInfo()
                    #archer.displayEqu()
                    self.aList.append(archer)

    def displayInfo(self):
        print("{:<25} {:<15} {:<15} {:<15} {:<15}".format("Name", "Age", "Ranking","Gender","Country"))
        for archer in self.aList:
            archer.displayInfo()

    def displayEquipStat(self):
        print("{:<25} {:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".format("Name", "Riser", "Limb", "Sightbar","Sightpin","Stabilizer","Tab"))
        for archer in self.aList:
            archer.displayEqu()
        


class Hoyt(Enum):
    GMX3=1
    Axia=2
    XD=3
    Xceed=4

class Winwin(Enum):
    ATFDX=1
    NSXP=2
    MXTXP=3
    ATFX=4
    MetaDX=5
    CX7=6
    S21=7
    NSG=8
    EZR=9
    ACSEL=10

class Fivics(Enum):
    ARGONX=1
    TITANNX=2
    saker1=3
    saker2=4
    saker3=5
    tenfix=6
    saker_plus=7

class MK(Enum):
    MKXG=1
    MKSD=2
    MKS=3
    X_core=4
    MX=5
    N3=6
    X_ON=7

class Gillo(Enum):
    GT25M=1
    GQ=2
    GT=3
    GX=4
    G1=5

class Axcel(Enum):
    XPpro=1
    XPvariable=2
    XP=3
    CarboFlax=4
    CarboFlaxpro=5
    Cube=6
    RXpro=7
    RXF=8
    Contour=9

class Shibuya(Enum):
    RCIII=1
    CPPro=2
    RCPro=3
    FiberOptic=4
    RecurveSightPin=5
    Vanquish=6
    Primus=7
    VERSA=8
    MZS=9
    Apex=10

class Fairweather(Enum):
    Modulus=1

class AAE(Enum):
    KSLGold=1
    Elite=2

class Conquest(Enum):
    Smacdown=1

class ULTRAVIEW(Enum):
    UV=1

class Doinker(Enum):
    Storm=1
    Champion=2

class Shrewd(Enum):
    S2=1
    RevX=2

class BeeStinger(Enum):
    PremierPlus=1
    Competitor=2
    Microhex=3

class Easton(Enum):
    Halcyon=1

class DeadCenter(Enum):
    IconX=1
