import openpyxl
import DataStruct

archerList=DataStruct.ArcherList()
file_path='./Equipment Statics.xlsx'
archerList.loadDatas(file_path)
#archerList.displayInfo()
archerList.displayEquipStat()
print(len(archerList.aList))

'''
workbook=openpyxl.load_workbook(file_path)
sheet_names=workbook.sheetnames
for s_name in sheet_names:
    sheet=workbook[s_name]
    print(f'Loading sheet: {s_name}')
    
    
    for row in range(2,sheet.max_row):
        temp=[]
        for cell in sheet[row]:
            temp.append(cell.value)
        if temp[5] != None and temp[0]!=None:
            archer=DataStruct.Archer(temp[0],temp[1],temp[2],temp[3],temp[4],temp[5],temp[6],temp[7],temp[8],temp[9],temp[10])
            archer.displayInfo()
            archer.displayEqu()
            archerlist.append(archer)
'''



        
        
    
    



'''
print(sheet_names)
sheet1=workbook.active

for row in sheet1.iter_rows(min_row=1,min_col=1,max_row=sheet1.max_row,max_col=sheet1.max_column,values_only=True):
#    print(row)
    pass

'''
